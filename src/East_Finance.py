import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font
import json
import time
import threading

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
}

class EastWealth(object):
    def __init__(self, file_name):
        self.xlsxname = file_name

    def request(self):
        self.url = 'https://www.eastmoney.com/'
        self.data = requests.get(self.url, headers=headers)
        self.data.encoding = "utf-8"
        self.soup = BeautifulSoup(self.data.text, "lxml")

        self.stock_url = 'http://stock.eastmoney.com/'
        self.stock_data = requests.get(self.stock_url, headers=headers)
        self.stock_data.encoding = "utf-8"
        self.stock_soup = BeautifulSoup(self.stock_data.text, "lxml")


    def getTopNew(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet("Ew")
        t_row = 1
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="东方财富")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        datalist = self.soup.find_all(class_='nlist')
        for Newslist in datalist:
            News = Newslist.find_all('a')
            for m_new in News:
                m_title = m_new.get_text()
                if len(m_title) <= 3:
                    break
                m_href = m_new['href']
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_href)
                t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except:
            print("EastWealth Save Error = getTopNew")


    def getStockNew(self):
        wb = load_workbook(self.xlsxname)
        sheet = wb.create_sheet("Ew")
        t_row = sheet.max_row + 2
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="股市聚焦")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1

        datalist = self.stock_soup.find_all(class_='card_body pt0 card_gsjj')
        for Newslist in datalist:
            News = Newslist.find_all('a')
            for m_new in News:
                print(m_new)


    def main(self, file_name):
        Ew = EastWealth(file_name)
        Ew.request()
        Ew.getTopNew()
        Ew.getStockNew()
